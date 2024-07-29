import time

from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.common.exceptions import (
    NoSuchElementException,
    ElementClickInterceptedException,
    StaleElementReferenceException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

chrome_options = Options()
# chrome_options.add_argument("--headless")  # 헤드리스 모드로 실행
# chrome_options.add_argument("--disable-gpu")  # GPU 사용 안함
chrome_options.add_argument("--no-sandbox")  # 샌드박스 사용 안함
# chrome_options.add_argument("--disable-dev-shm-usage")  # /dev/shm 사용 안함
chrome_options.add_argument("--enable-precise-memory-info")
chrome_options.add_argument("--disable-popup-blocking")
chrome_options.add_argument("--disable-default-apps")
chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--disable-infobars")
chrome_options.add_argument("--disable-notifications")
chrome_options.add_argument("--lang=ko")  # 브라우저 언어 설정을 한국어로 변경
chrome_options.page_load_strategy = "eager"  # 페이지 로딩 최적화
# 불필요한 리소스 로딩 차단
prefs = {
    "profile.managed_default_content_settings": {
        "images": 2,  # 이미지 로딩 차단
    }
}
chrome_options.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome("./chromedriver", options=chrome_options)
driver.set_script_timeout(300)
wait = WebDriverWait(driver, 300)

driver.get("https://imweb.me/login")
driver.implicitly_wait(1)

id_input = driver.find_element(By.XPATH, '//*[@id="io-email-field-input-2"]')
driver.implicitly_wait(1)
id_input.send_keys("kcs15985@naver.com")
driver.implicitly_wait(1)

pw_input = driver.find_element(By.XPATH, '//*[@id="io-password-field-input-3"]')
driver.implicitly_wait(1)
pw_input.send_keys("Dnl1vhr2@")
driver.implicitly_wait(1)

login_btn = driver.find_element(
    By.XPATH, '//*[@id="snowfall"]/main/io-login-form/io-login-form-container/form/fieldset/io-button/button'
)
driver.implicitly_wait(1)
login_btn.click()
driver.implicitly_wait(1)

driver.get("https://imweb.me/expert")
driver.implicitly_wait(1)
select_btn = wait.until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/div/main/nav/nav/ul/div/div[2]/div/button'))
)
select_btn.click()
like_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[5]/ul/li[3]/button")))
like_btn.click()

driver.execute_script(f"window.scrollBy(0, 100);")
driver.implicitly_wait(1)
print("login")

company_datas = []

for i in range(1, 1000000):
    print(i)
    try:
        data = {
            "company_name": None,
            "email": None,
            "phone": None,
        }

        try:
            content = wait.until(
                EC.element_to_be_clickable((By.XPATH, f'//*[@id="__next"]/div/main/div[3]/div/div/section[{i}]/div'))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", content)
            driver.execute_script("window.scrollBy(0, -60);")

            company_name = driver.find_element(
                By.XPATH, f'//*[@id="__next"]/div/main/div[3]/div/div/section[{i}]/label/div[1]/a/p'
            )

            company_exists = False
            for company in company_datas:
                if company["company_name"] == company_name.text:
                    company_exists = True
                    break

            if company_exists:
                continue

            data["company_name"] = company_name.text

            print(data["company_name"])
            content.click()
        except (ElementClickInterceptedException, StaleElementReferenceException) as e:
            print("fail : ", e)
            continue

        try:
            inquiry_btn = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, '//*[@id="customModal"]/section/div/div/div[2]/div/div/aside/div/div[2]/button')
                )
            )
            inquiry_btn.click()
        except (ElementClickInterceptedException, StaleElementReferenceException) as e:
            print("retrying inquiry button click after failure: ", e)
            time.sleep(5)
            inquiry_btn = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, '//*[@id="customModal"]/section/div/div/div[2]/div/div/aside/div/div[2]/button')
                )
            )
            inquiry_btn.click()

        try:
            company_name_element = driver.find_element(
                By.XPATH, "/html/body/div[5]/div/section/div/div[2]/div/div[1]/div[1]/p"
            )

            company_name_text = company_name_element.text
            data["company_name"] = company_name_text
        except NoSuchElementException:
            pass
        driver.implicitly_wait(1)
        try:
            introduction_element = driver.find_element(
                By.XPATH, "/html/body/div[5]/div/section/div/div[2]/div/div[1]/article/p"
            )

            introduction_text = introduction_element.text
            data["introduction"] = introduction_text
        except NoSuchElementException:
            pass
        driver.implicitly_wait(1)
        try:
            email_element = driver.find_element(
                By.XPATH, "/html/body/div[5]/div/section/div/div[2]/div/div[1]/div[2]/div[1]/div[2]/p[2]"
            )

            email_text = email_element.text
            data["email"] = email_text
        except NoSuchElementException:
            pass
        driver.implicitly_wait(1)
        try:
            phone_element = driver.find_element(
                By.XPATH, "/html/body/div[5]/div/section/div/div[2]/div/div[1]/div[2]/div[2]/div[2]/p[2]"
            )

            phone_text = phone_element.text
            data["phone"] = phone_text
        except NoSuchElementException:
            pass
        print("information")
        driver.implicitly_wait(1)
        try:
            inquiry_close_btn = wait.until(
                EC.element_to_be_clickable((By.XPATH, "/html/body/div[5]/div/section/div/div[1]/button"))
            )
            inquiry_close_btn.click()
        except (ElementClickInterceptedException, StaleElementReferenceException) as e:
            print("retrying inquiry close button click after failure: ", e)
            time.sleep(5)
            inquiry_close_btn = wait.until(
                EC.element_to_be_clickable((By.XPATH, "/html/body/div[5]/div/section/div/div[1]/button"))
            )
            inquiry_close_btn.click()

        try:
            content_close_btn = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, '//*[@id="customModal"]/section/div/div/div[2]/div/div/div/div/button/button')
                )
            )
            content_close_btn.click()
        except (ElementClickInterceptedException, StaleElementReferenceException) as e:
            print("retrying content close button click after failure: ", e)
            time.sleep(5)
            content_close_btn = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, '//*[@id="customModal"]/section/div/div/div[2]/div/div/div/div/button/button')
                )
            )
            content_close_btn.click()

        company_datas.append(data)
        print(data["email"])

    except NoSuchElementException as e:
        print("fail : ", e)
        break
    except KeyboardInterrupt:
        break

driver.quit()


wb = Workbook()

for sheet in wb.sheetnames:
    wb.remove(wb[sheet])


ws = wb.create_sheet(title="아임웹 전문가")
ws.row_dimensions[1].height = 10
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 40
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 100

ws["A1"] = "No"
ws["B1"] = "화사 이름"
ws["C1"] = "이메일"
ws["D1"] = "휴대폰"
ws["E1"] = "소개글"


title_cells = [ws["A1"], ws["B1"], ws["C1"], ws["D1"], ws["E1"]]


row = 2

for idx, data in enumerate(company_datas):
    ws[f"A{row}"] = idx + 1
    ws[f"B{row}"] = data["company_name"]
    ws[f"C{row}"] = data["email"]
    ws[f"D{row}"] = data["phone"]
    ws[f"E{row}"] = data["introduction"]

    row += 1


wb.save("imweb_company_data.xlsx")
